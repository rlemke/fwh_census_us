"""External county-indicator sources: County Health Rankings + HUD PIT.

These fill national county metrics no ACS table carries:

- **Homicides** (``chr_homicide``) — CHR 2025 measure v015: CDC NVSS death
  records, deaths per 100,000, pooled over multiple years to reduce
  small-county suppression (still ~1/3 of counties suppressed → None).
- **Violent crime** (``chr_violent_crime``) — CHR 2022 measure v043: FBI UCR
  offenses per 100,000, 2014/2016 data — the LAST comparable national county
  series (agency reporting broke in the UCR→NIBRS transition).
- **Jobless time series** — CHR trends "Unemployment rate": BLS-derived annual
  county unemployment 2002-2022 (fraction → percent).
- **Homeless time series** — HUD Point-in-Time "Overall Homeless" by
  Continuum of Care (one sheet per year, 2007+), apportioned to counties by
  each county's population share of its CoC (Byrne HUD-CoC crosswalk), as a
  rate per 10,000 residents. An explicit APPROXIMATION: CoCs are not county
  aggregates, the crosswalk is 2017-vintage, and the population denominator
  is current-ACS. Labeled as such on the map.

Every fetch caches the raw file under cache/indicators/ (cstore-backed) and
each normalizer writes a CSV shaped like the ACS downloader's output
(GEOID "0500000US<fips>", NAME, value columns) so the map builders can merge
them exactly like ACS files. Parsers are separated from fetchers so tests run
offline on synthetic files.
"""

from __future__ import annotations

import csv
import logging
import re
from datetime import UTC, datetime
from typing import Any

try:
    import requests

    HAS_REQUESTS = True
except ImportError:
    HAS_REQUESTS = False

from census_us.tools._lib import storage as cstore

logger = logging.getLogger(__name__)

CHR_BASE = "https://www.countyhealthrankings.org/sites/default/files/media/document"
CHR_HOMICIDE_URL = f"{CHR_BASE}/analytic_data2025_v2.csv"  # v015
CHR_CRIME_URL = f"{CHR_BASE}/analytic_data2022.csv"  # v043 (last release with it)
CHR_TRENDS_URL = f"{CHR_BASE}/chr_trends_csv_2024.csv"  # Unemployment rate 2002-2022
PIT_URL = "https://www.huduser.gov/portal/sites/default/files/xls/2007-2024-PIT-Counts-by-CoC.xlsb"
COC_XWALK_URL = (
    "https://raw.githubusercontent.com/tomhbyrne/HUD-CoC-Geography-Crosswalk/"
    "master/output/county_coc_match.csv"
)

# huduser.gov sits behind a bot filter that serves an HTML challenge to
# default client UAs; a browser UA gets the real file.
_BROWSER_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0 Safari/537.36"
    ),
    "Accept": "*/*",
}

_COC_RE = re.compile(r"^[A-Z]{2}-\d{3}$")


def _fetch(url: str, dest: str, *, browser: bool = False, min_bytes: int = 10000) -> str:
    """Download url → cstore dest (cache-first). Returns the dest path."""
    if cstore.exists(dest) and cstore.size(dest) >= min_bytes:
        logger.info("indicator cache hit: %s", dest)
        return dest
    if not HAS_REQUESTS:
        raise RuntimeError("requests library required for downloads")
    logger.info("Fetching %s -> %s", url, dest)
    resp = requests.get(
        url, timeout=300, stream=True, headers=_BROWSER_HEADERS if browser else None
    )
    resp.raise_for_status()
    size = 0
    with cstore.open_write(dest, "wb") as f:
        for chunk in resp.iter_content(chunk_size=65536):
            f.write(chunk)
            size += len(chunk)
    if size < min_bytes:
        raise RuntimeError(
            f"Suspiciously small download ({size}B) from {url} — likely a bot-challenge page"
        )
    return dest


def _file_info(path: str, url: str, was_cached: bool) -> dict[str, Any]:
    return {
        "url": url,
        "path": path,
        "date": datetime.now(UTC).isoformat(),
        "size": cstore.size(path),
        "wasInCache": was_cached,
    }


# ---------------------------------------------------------------------------
# County Health Rankings — homicide + violent crime snapshot indicators.
# ---------------------------------------------------------------------------


def parse_chr_measure(
    path: str, measure_col: str, *, scale: float = 1.0, round_nd: int = 1
) -> dict[str, tuple[float, str]]:
    """{5-digit fips: (rawvalue*scale, "County, ST")} from a CHR analytic CSV.

    Row 1 is human-readable headers, row 2 machine names — DictReader keys on
    row 1's cells only if we skip it, so read via the machine-name row.
    ``scale`` converts fraction measures (obesity 0.384) to percent.
    """
    out: dict[str, tuple[float, str]] = {}
    with cstore.open_read(path, newline="") as f:
        reader = csv.reader(f)
        next(reader)  # human header
        machine = next(reader)
        idx = {c: i for i, c in enumerate(machine)}
        for col in ("fipscode", "county", "state", "countycode"):
            if col not in idx:
                raise ValueError(f"CHR file missing column {col}")
        if measure_col not in idx:
            raise ValueError(f"CHR file missing measure column {measure_col}")
        for row in reader:
            if len(row) <= idx[measure_col]:
                continue
            if (row[idx["countycode"]] or "000") == "000":
                continue  # US + state rows
            fips = (row[idx["fipscode"]] or "").zfill(5)
            try:
                val = float(row[idx[measure_col]])
            except (TypeError, ValueError):
                continue
            name = f"{row[idx['county']]}, {row[idx['state']]}"
            out[fips] = (round(val * scale, round_nd), name)
    return out


def build_chr_indicators_csv() -> dict[str, Any]:
    """Fetch CHR releases and write the normalized homicide+crime CSV."""
    cache = cstore.join(cstore.cache_root(), "indicators")
    homicide_raw = _fetch(CHR_HOMICIDE_URL, cstore.join(cache, "chr_analytic_2025.csv"))
    crime_raw = _fetch(CHR_CRIME_URL, cstore.join(cache, "chr_analytic_2022.csv"))
    homicide = parse_chr_measure(homicide_raw, "v015_rawvalue")
    crime = parse_chr_measure(crime_raw, "v043_rawvalue")

    dest = cstore.join(cache, "chr_indicators.csv")
    all_fips = sorted(set(homicide) | set(crime))
    with cstore.open_write(dest, "w", newline="") as f:
        wr = csv.writer(f)
        wr.writerow(["GEOID", "NAME", "chr_homicide", "chr_violent_crime"])
        for fips in all_fips:
            name = (homicide.get(fips) or crime.get(fips))[1]
            wr.writerow([
                f"0500000US{fips}", name,
                homicide[fips][0] if fips in homicide else "",
                crime[fips][0] if fips in crime else "",
            ])
    logger.info(
        "CHR indicators: %d counties (homicide %d, crime %d) -> %s",
        len(all_fips), len(homicide), len(crime), dest,
    )
    return _file_info(dest, CHR_HOMICIDE_URL, False)


# ---------------------------------------------------------------------------
# County Health Rankings trends — annual county unemployment 2002-2022.
# ---------------------------------------------------------------------------


def parse_chr_jobless_trends(path: str) -> dict[str, dict[int, float]]:
    """{fips: {year: pct}} from the CHR trends CSV ("Unemployment rate")."""
    out: dict[str, dict[int, float]] = {}
    with cstore.open_read(path, newline="") as f:
        for row in csv.DictReader(f):
            if row.get("measurename") != "Unemployment rate":
                continue
            county = (row.get("countycode") or "000").strip()
            state = (row.get("statecode") or "00").strip()
            if county == "000" or state == "00":
                continue
            try:
                year = int((row.get("yearspan") or "").strip())
                val = float(row.get("rawvalue", "").replace(",", "").strip('"'))
            except (TypeError, ValueError):
                continue
            out.setdefault(state.zfill(2) + county.zfill(3), {})[year] = round(val * 100, 2)
    return out


def build_chr_jobless_ts_csv() -> dict[str, Any]:
    """Fetch the CHR trends CSV and write the wide jobless time-series CSV."""
    cache = cstore.join(cstore.cache_root(), "indicators")
    raw = _fetch(CHR_TRENDS_URL, cstore.join(cache, "chr_trends_2024.csv"), min_bytes=1000000)
    series = parse_chr_jobless_trends(raw)
    years = sorted({y for by_year in series.values() for y in by_year})
    dest = cstore.join(cache, "chr_jobless_ts.csv")
    with cstore.open_write(dest, "w", newline="") as f:
        wr = csv.writer(f)
        wr.writerow(["GEOID", "NAME"] + [f"y{y}" for y in years])
        for fips in sorted(series):
            wr.writerow(
                [f"0500000US{fips}", ""]
                + [series[fips].get(y, "") for y in years]
            )
    logger.info("CHR jobless trends: %d counties x %d years -> %s", len(series), len(years), dest)
    return _file_info(dest, CHR_TRENDS_URL, False)


# ---------------------------------------------------------------------------
# HUD PIT — homeless per 10k residents, apportioned CoC → county.
# ---------------------------------------------------------------------------


def parse_coc_crosswalk(path: str) -> dict[str, list[tuple[str, float]]]:
    """{fips: [(coc_number, pct_of_county_pop_in_coc 0..1), …]}."""
    out: dict[str, list[tuple[str, float]]] = {}
    # cp1252 bytes in CoC names ("Prince George's") — decode accordingly.
    local = cstore.localize(path)
    with open(local, newline="", encoding="cp1252") as f:
        for row in csv.DictReader(f):
            fips = (row.get("county_fips") or "").zfill(5)
            coc = (row.get("coc_number") or "").strip()
            if len(fips) != 5 or not _COC_RE.match(coc):
                continue
            try:
                pct = float(row.get("pct_cnty_pop_coc") or 0) / 100.0
            except (TypeError, ValueError):
                continue
            if pct > 0:
                out.setdefault(fips, []).append((coc, pct))
    return out


def parse_pit_workbook(path: str) -> dict[int, dict[str, float]]:
    """{year: {coc_number: overall homeless}} from the PIT xlsb workbook."""
    from pyxlsb import open_workbook

    local = cstore.localize(path)
    out: dict[int, dict[str, float]] = {}
    with open_workbook(local) as wb:
        for sheet in wb.sheets:
            if not (sheet.isdigit() and 2007 <= int(sheet) <= 2100):
                continue
            year = int(sheet)
            with wb.get_sheet(sheet) as ws:
                coc_i = overall_i = None
                header_tries = 0
                for row in ws.rows():
                    vals = [c.v for c in row]
                    if coc_i is None or overall_i is None:
                        # scan the first rows for the header (usually row 1)
                        header_tries += 1
                        if header_tries > 5:
                            break
                        header = [str(v or "").strip() for v in vals]
                        for i, h in enumerate(header):
                            if h == "CoC Number":
                                coc_i = i
                            # "Overall Homeless" or "Overall Homeless, <year>"
                            if h in ("Overall Homeless", f"Overall Homeless, {year}"):
                                overall_i = i
                        continue
                    coc = str(vals[coc_i] or "").strip() if len(vals) > coc_i else ""
                    if not _COC_RE.match(coc):
                        continue
                    try:
                        out.setdefault(year, {})[coc] = float(vals[overall_i])
                    except (TypeError, ValueError, IndexError):
                        continue
    return out


def apportion_homeless(
    pit: dict[int, dict[str, float]],
    xwalk: dict[str, list[tuple[str, float]]],
    county_pop: dict[str, float],
) -> dict[str, dict[int, float]]:
    """{fips: {year: rate per 10k}} — CoC totals split by county pop share."""
    # county share of each CoC's population
    coc_pop: dict[str, float] = {}
    for fips, links in xwalk.items():
        pop = county_pop.get(fips)
        if not pop:
            continue
        for coc, pct in links:
            coc_pop[coc] = coc_pop.get(coc, 0.0) + pop * pct
    out: dict[str, dict[int, float]] = {}
    for fips, links in xwalk.items():
        pop = county_pop.get(fips)
        if not pop:
            continue
        for year, totals in pit.items():
            homeless = 0.0
            seen = False
            for coc, pct in links:
                total = totals.get(coc)
                if total is None or not coc_pop.get(coc):
                    continue
                homeless += total * (pop * pct) / coc_pop[coc]
                seen = True
            if seen:
                out.setdefault(fips, {})[year] = round(homeless / pop * 10000, 2)
    return out


def build_homeless_ts_csv(acs_path: str) -> dict[str, Any]:
    """Fetch PIT + crosswalk, apportion, and write the wide homeless CSV.

    ``acs_path`` is a national county ACS pull carrying B01003_001E (total
    population) — the apportionment weight and the rate denominator.
    """
    cache = cstore.join(cstore.cache_root(), "indicators")
    pit_raw = _fetch(
        PIT_URL, cstore.join(cache, "pit_counts_by_coc.xlsb"),
        browser=True, min_bytes=1000000,
    )
    xwalk_raw = _fetch(COC_XWALK_URL, cstore.join(cache, "county_coc_match.csv"))

    county_pop: dict[str, float] = {}
    with cstore.open_read(acs_path, newline="") as f:
        for row in csv.DictReader(f):
            fips = (row.get("GEOID", "") or "").rsplit("US", 1)[-1]
            try:
                pop = float(row.get("B01003_001E", ""))
            except (TypeError, ValueError):
                continue
            if len(fips) == 5 and pop > 0:
                county_pop[fips] = pop

    pit = parse_pit_workbook(pit_raw)
    xwalk = parse_coc_crosswalk(xwalk_raw)
    rates = apportion_homeless(pit, xwalk, county_pop)
    years = sorted({y for by_year in rates.values() for y in by_year})

    dest = cstore.join(cache, "hud_homeless_ts.csv")
    with cstore.open_write(dest, "w", newline="") as f:
        wr = csv.writer(f)
        wr.writerow(["GEOID", "NAME"] + [f"y{y}" for y in years])
        for fips in sorted(rates):
            wr.writerow(
                [f"0500000US{fips}", ""]
                + [rates[fips].get(y, "") for y in years]
            )
    logger.info(
        "HUD homeless: %d CoC-years, %d counties x %d years -> %s",
        sum(len(v) for v in pit.values()), len(rates), len(years), dest,
    )
    return _file_info(dest, PIT_URL, False)


# ---------------------------------------------------------------------------
# CHR RELEASE SERIES — one measure across every annual release (2010-2025).
# ---------------------------------------------------------------------------

# Older releases live at a different path; 2025 is the _v2 revision.
def _chr_release_url(release: int) -> str:
    if release <= 2018:
        return ("https://www.countyhealthrankings.org/sites/default/files/"
                f"analytic_data{release}.csv")
    if release == 2025:
        return f"{CHR_BASE}/analytic_data2025_v2.csv"
    return f"{CHR_BASE}/analytic_data{release}.csv"


# measure key -> (CHR machine column, value scale, data-year offset from the
# release year, first release carrying it, output column). The offset labels
# each release's frame by its approximate DATA year: obesity (v011) is a
# BRFSS-modeled single year ~3 years before release; life expectancy (v147)
# is a 3-year NVSS window roughly centered 3 years before release (windows
# overlap — disclosed in the map's About text).
CHR_SERIES = {
    "obesity": ("v011_rawvalue", 100.0, -3, 2010, "chr_obesity"),
    "life_expectancy": ("v147_rawvalue", 1.0, -3, 2019, "chr_life_expectancy"),
    "smoking": ("v009_rawvalue", 100.0, -3, 2010, "chr_smoking"),
    # v060 first appears in the 2013 release (earlier releases auto-skip)
    "diabetes": ("v060_rawvalue", 100.0, -3, 2013, "chr_diabetes"),
}

CHR_LATEST_RELEASE = 2025


def build_chr_measure_series_csv(measure: str) -> dict[str, Any]:
    """Fetch every CHR release carrying ``measure`` and write a wide time CSV
    (GEOID, NAME, y<data_year>…). Releases missing the column are skipped."""
    if measure not in CHR_SERIES:
        raise ValueError(f"Unknown CHR series measure: {measure}. Known: {sorted(CHR_SERIES)}")
    col, scale, offset, first_release, _out_col = CHR_SERIES[measure]
    cache = cstore.join(cstore.cache_root(), "indicators")

    series: dict[str, dict[int, float]] = {}
    names: dict[str, str] = {}
    for release in range(first_release, CHR_LATEST_RELEASE + 1):
        dest = cstore.join(cache, f"chr_release_{release}.csv")
        raw = _fetch(_chr_release_url(release), dest, min_bytes=1000000)
        try:
            vals = parse_chr_measure(raw, col, scale=scale)
        except ValueError:
            logger.info("CHR %d lacks %s — skipped", release, col)
            continue
        year = release + offset
        for fips, (v, name) in vals.items():
            series.setdefault(fips, {})[year] = v
            names.setdefault(fips, name)

    years = sorted({y for by_year in series.values() for y in by_year})
    if not years:
        raise RuntimeError(f"No CHR release carried {col}")
    dest = cstore.join(cache, f"chr_{measure}_ts.csv")
    with cstore.open_write(dest, "w", newline="") as f:
        wr = csv.writer(f)
        wr.writerow(["GEOID", "NAME"] + [f"y{y}" for y in years])
        for fips in sorted(series):
            wr.writerow(
                [f"0500000US{fips}", names.get(fips, "")]
                + [series[fips].get(y, "") for y in years]
            )
    logger.info(
        "CHR %s series: %d counties x %d frames (%d-%d) -> %s",
        measure, len(series), len(years), years[0], years[-1], dest,
    )
    return _file_info(dest, _chr_release_url(CHR_LATEST_RELEASE), False)


# ---------------------------------------------------------------------------
# CDC heart-disease mortality trends (annual, 1999-2019) + NCI cancer snapshot.
# ---------------------------------------------------------------------------

HEART_TRENDS_URL = (
    "https://data.cdc.gov/resource/7b9s-s8ck.csv"
    "?$limit=200000&$select=year,locationid,data_value"
    "&geographiclevel=County&stratification2=Overall&stratification3=Overall"
)
CANCER_SCP_URL = (
    "https://statecancerprofiles.cancer.gov/deathrates/index.php"
    "?stateFIPS=00&areatype=county&cancer=001&race=00&sex=0&age=001&year=0"
    "&type=death&sortVariableName=rate&sortOrder=default&output=1"
)


def parse_heart_trends_csv(path: str) -> dict[str, dict[int, float]]:
    """{fips: {year: rate}} from the filtered Socrata CSV (year, locationid,
    data_value). Pooled-window rows ("1999 - 2010") and suppressed values
    ("NA") are skipped."""
    out: dict[str, dict[int, float]] = {}
    with cstore.open_read(path, newline="") as f:
        for row in csv.DictReader(f):
            y = (row.get("year") or "").strip()
            fips = (row.get("locationid") or "").strip().zfill(5)
            if not y.isdigit() or len(fips) != 5:
                continue
            try:
                val = float(row.get("data_value", ""))
            except (TypeError, ValueError):
                continue
            out[fips] = out.get(fips, {})
            out[fips][int(y)] = round(val, 1)
    return out


def build_heart_disease_ts_csv(
    topic: str = "All heart disease", age: str = "Ages 35-64 years"
) -> dict[str, Any]:
    """CDC 'Rates and Trends in Heart Disease and Stroke Mortality' →
    wide annual time CSV (age-standardized, spatiotemporally smoothed,
    per 100k, NVSS). One Socrata pull for the Overall/Overall slice."""
    from urllib.parse import quote

    url = f"{HEART_TRENDS_URL}&topic={quote(topic)}&stratification1={quote(age)}"
    cache = cstore.join(cstore.cache_root(), "indicators")
    slug = f"{topic}_{age}".lower().replace(" ", "_").replace("/", "-")[:60]
    raw = _fetch(url, cstore.join(cache, f"cdc_heart_{slug}.csv"), min_bytes=100000)
    series = parse_heart_trends_csv(raw)
    years = sorted({y for by_year in series.values() for y in by_year})
    if not years:
        raise RuntimeError(f"No heart-disease rows for topic={topic!r} age={age!r}")
    dest = cstore.join(cache, "cdc_heart_disease_ts.csv")
    with cstore.open_write(dest, "w", newline="") as f:
        wr = csv.writer(f)
        wr.writerow(["GEOID", "NAME"] + [f"y{y}" for y in years])
        for fips in sorted(series):
            wr.writerow(
                [f"0500000US{fips}", ""]
                + [series[fips].get(y, "") for y in years]
            )
    logger.info(
        "CDC heart trends: %d counties x %d years -> %s", len(series), len(years), dest
    )
    return _file_info(dest, url, False)


def parse_scp_deathrates(path: str) -> dict[str, tuple[float, str]]:
    """{fips: (age-adjusted rate, county name)} from a State Cancer Profiles
    death-rate export: a preamble, then a County,FIPS,… header, then data rows,
    then footnotes. Rates carry footnote whitespace ("412.2 ")."""
    out: dict[str, tuple[float, str]] = {}
    with cstore.open_read(path, newline="") as f:
        reader = csv.reader(f)
        header = None
        rate_i = None
        for row in reader:
            if header is None:
                if row and row[0].strip() == "County" and "FIPS" in [c.strip() for c in row[:3]]:
                    header = [c.strip() for c in row]
                    rate_i = next(
                        (i for i, c in enumerate(header) if c.startswith("Age-Adjusted Death Rate")),
                        None,
                    )
                continue
            if rate_i is None or len(row) <= rate_i:
                continue
            fips = (row[1] or "").strip()
            if len(fips) != 5 or not fips.isdigit() or fips == "00000":
                continue
            try:
                val = float((row[rate_i] or "").strip())
            except (TypeError, ValueError):
                continue
            out[fips] = (round(val, 1), (row[0] or "").strip())
    return out


def build_cancer_mortality_csv() -> dict[str, Any]:
    """NCI State Cancer Profiles all-counties, all-site cancer death rates
    (age-adjusted per 100k, latest 5-year window) → normalized indicator CSV."""
    cache = cstore.join(cstore.cache_root(), "indicators")
    raw = _fetch(
        CANCER_SCP_URL, cstore.join(cache, "scp_cancer_deathrates.csv"),
        browser=True, min_bytes=100000,
    )
    vals = parse_scp_deathrates(raw)
    dest = cstore.join(cache, "scp_cancer_mortality.csv")
    with cstore.open_write(dest, "w", newline="") as f:
        wr = csv.writer(f)
        wr.writerow(["GEOID", "NAME", "scp_cancer_mortality"])
        for fips in sorted(vals):
            v, name = vals[fips]
            wr.writerow([f"0500000US{fips}", name, v])
    logger.info("SCP cancer mortality: %d counties -> %s", len(vals), dest)
    return _file_info(dest, CANCER_SCP_URL, False)


# ---------------------------------------------------------------------------
# NCHS drug-poisoning (overdose) mortality — annual county rates 2003-2021.
# ---------------------------------------------------------------------------

DRUG_OD_URL = (
    "https://data.cdc.gov/resource/rpvx-m2md.csv"
    "?$limit=200000&$select=year,fips,model_based_death_rate"
)


def parse_drug_od_csv(path: str) -> dict[str, dict[int, float]]:
    """{fips: {year: rate}} — Socrata strips leading zeros from fips."""
    out: dict[str, dict[int, float]] = {}
    with cstore.open_read(path, newline="") as f:
        for row in csv.DictReader(f):
            y = (row.get("year") or "").strip()
            fips = (row.get("fips") or "").strip().zfill(5)
            if not y.isdigit() or len(fips) != 5:
                continue
            try:
                val = float(row.get("model_based_death_rate", ""))
            except (TypeError, ValueError):
                continue
            out.setdefault(fips, {})[int(y)] = round(val, 1)
    return out


def build_drug_overdose_ts_csv() -> dict[str, Any]:
    """NCHS model-based county drug-poisoning death rates (annual 2003-2021,
    per 100k) → wide time CSV. One Socrata pull."""
    cache = cstore.join(cstore.cache_root(), "indicators")
    raw = _fetch(DRUG_OD_URL, cstore.join(cache, "nchs_drug_od.csv"), min_bytes=500000)
    series = parse_drug_od_csv(raw)
    years = sorted({y for by_year in series.values() for y in by_year})
    if not years:
        raise RuntimeError("No drug-overdose rows parsed")
    dest = cstore.join(cache, "nchs_drug_od_ts.csv")
    with cstore.open_write(dest, "w", newline="") as f:
        wr = csv.writer(f)
        wr.writerow(["GEOID", "NAME"] + [f"y{y}" for y in years])
        for fips in sorted(series):
            wr.writerow(
                [f"0500000US{fips}", ""]
                + [series[fips].get(y, "") for y in years]
            )
    logger.info("NCHS drug OD: %d counties x %d years -> %s", len(series), len(years), dest)
    return _file_info(dest, DRUG_OD_URL, False)


# ---------------------------------------------------------------------------
# CDC injury surveillance — county suicide rates, annual 2019-2024.
# ---------------------------------------------------------------------------

SUICIDE_URL = (
    "https://data.cdc.gov/resource/psx4-wq38.csv"
    "?$limit=50000&$select=period,geoid,rate&intent=All_Suicide"
)


def parse_suicide_csv(path: str) -> dict[str, dict[int, float]]:
    """{fips: {year: rate}} — non-annual periods (TTM) are skipped."""
    out: dict[str, dict[int, float]] = {}
    with cstore.open_read(path, newline="") as f:
        for row in csv.DictReader(f):
            y = (row.get("period") or "").strip()
            fips = (row.get("geoid") or "").strip().zfill(5)
            if not y.isdigit() or len(fips) != 5:
                continue
            try:
                val = float(row.get("rate", ""))
            except (TypeError, ValueError):
                continue
            out.setdefault(fips, {})[int(y)] = round(val, 1)
    return out


def build_suicide_ts_csv() -> dict[str, Any]:
    """CDC injury-surveillance county suicide rates (annual 2019-2024, per
    100k; small-count rates are flagged unstable by CDC but published) →
    wide time CSV. One data.cdc.gov pull."""
    cache = cstore.join(cstore.cache_root(), "indicators")
    raw = _fetch(SUICIDE_URL, cstore.join(cache, "cdc_suicide.csv"), min_bytes=100000)
    series = parse_suicide_csv(raw)
    years = sorted({y for by_year in series.values() for y in by_year})
    if not years:
        raise RuntimeError("No suicide rows parsed")
    dest = cstore.join(cache, "cdc_suicide_ts.csv")
    with cstore.open_write(dest, "w", newline="") as f:
        wr = csv.writer(f)
        wr.writerow(["GEOID", "NAME"] + [f"y{y}" for y in years])
        for fips in sorted(series):
            wr.writerow(
                [f"0500000US{fips}", ""]
                + [series[fips].get(y, "") for y in years]
            )
    logger.info("CDC suicide: %d counties x %d years -> %s", len(series), len(years), dest)
    return _file_info(dest, SUICIDE_URL, False)


# ---------------------------------------------------------------------------
# Pew Research — unauthorized immigrant population by STATE, 1990-2023.
# ---------------------------------------------------------------------------

PEW_STATES_URL = (
    "https://www.pewresearch.org/wp-content/uploads/sites/20/2025/08/"
    "RE_2025.08.21_Unauthorized-immigrants_detailed-tables_state-trends.xlsx"
)


def parse_pew_state_trends(path: str) -> dict[str, dict[int, float]]:
    """{2-digit state fips: {year: estimate}} from Pew's state-trends XLSX.

    Header row 4 carries year columns ("'23", "'90") for the estimates block,
    then a spacer and the margin-of-error block (ignored — we stop at the
    first empty header after the estimates begin). Rows are state names;
    "U.S. total" is skipped. Years '90-'99 → 19xx, else 20xx (no 2020 —
    pandemic survey disruption).
    """
    import openpyxl

    from census_us.tools._lib.maps import STATE_FIPS

    local = cstore.localize(path)
    wb = openpyxl.load_workbook(local, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    header = rows[3]
    year_cols: list[tuple[int, int]] = []  # (col index, year)
    started = False
    for i, h in enumerate(header):
        hs = str(h or "").strip()
        if hs.startswith("'") and hs[1:].isdigit():
            yy = int(hs[1:])
            year_cols.append((i, 1900 + yy if yy >= 90 else 2000 + yy))
            started = True
        elif started:
            break  # spacer before the margin-of-error block

    out: dict[str, dict[int, float]] = {}
    for row in rows[4:]:
        name = str(row[1] or "").strip() if len(row) > 1 else ""
        fips = STATE_FIPS.get(name)
        if not fips:
            continue
        for i, year in year_cols:
            try:
                val = float(row[i])
            except (TypeError, ValueError, IndexError):
                continue
            out.setdefault(fips, {})[year] = val
    return out


def build_unauthorized_ts_csv() -> dict[str, Any]:
    """Pew unauthorized-population state series (1990-2023, ~24 vintages) →
    wide STATE-level time CSV (GEOID 0400000US<st>)."""
    cache = cstore.join(cstore.cache_root(), "indicators")
    raw = _fetch(
        PEW_STATES_URL, cstore.join(cache, "pew_state_trends.xlsx"),
        browser=True, min_bytes=10000,
    )
    series = parse_pew_state_trends(raw)
    years = sorted({y for by_year in series.values() for y in by_year})
    if not years:
        raise RuntimeError("No Pew state rows parsed")
    from census_us.tools._lib.maps import _FIPS_NAME

    dest = cstore.join(cache, "pew_unauthorized_ts.csv")
    with cstore.open_write(dest, "w", newline="") as f:
        wr = csv.writer(f)
        wr.writerow(["GEOID", "NAME"] + [f"y{y}" for y in years])
        for fips in sorted(series):
            wr.writerow(
                [f"0400000US{fips}", _FIPS_NAME.get(fips, "")]
                + [series[fips].get(y, "") for y in years]
            )
    logger.info(
        "Pew unauthorized: %d states x %d vintages (%d-%d) -> %s",
        len(series), len(years), years[0], years[-1], dest,
    )
    return _file_info(dest, PEW_STATES_URL, False)


# ---------------------------------------------------------------------------
# Zillow ZHVI — median home value by county, annual (June) 2000-present.
# ---------------------------------------------------------------------------

ZHVI_URL = (
    "https://files.zillowstatic.com/research/public_csvs/zhvi/"
    "County_zhvi_uc_sfrcondo_tier_0.33_0.67_sm_sa_month.csv"
)


def parse_zhvi_csv(path: str) -> dict[str, dict[int, float]]:
    """{fips: {year: June ZHVI}} from Zillow's county workbook (monthly
    columns; one mid-year frame per year keeps map payloads sane)."""
    out: dict[str, dict[int, float]] = {}
    with cstore.open_read(path, newline="") as f:
        reader = csv.DictReader(f)
        june_cols = [c for c in (reader.fieldnames or []) if c.endswith("-06-30")]
        for row in reader:
            if (row.get("RegionType") or "") != "county":
                continue
            fips = (row.get("StateCodeFIPS") or "").zfill(2) + (
                row.get("MunicipalCodeFIPS") or ""
            ).zfill(3)
            if len(fips) != 5:
                continue
            for c in june_cols:
                try:
                    val = float(row.get(c, ""))
                except (TypeError, ValueError):
                    continue
                out.setdefault(fips, {})[int(c[:4])] = round(val)
    return out


def build_home_value_ts_csv() -> dict[str, Any]:
    """Zillow ZHVI county series → wide annual (June) time CSV, 2000-present.
    ZHVI = smoothed, seasonally-adjusted typical home value (35th-65th
    percentile, single-family + condo). Data provided by Zillow."""
    cache = cstore.join(cstore.cache_root(), "indicators")
    raw = _fetch(ZHVI_URL, cstore.join(cache, "zillow_zhvi_county.csv"), min_bytes=1000000)
    series = parse_zhvi_csv(raw)
    years = sorted({y for by_year in series.values() for y in by_year})
    if not years:
        raise RuntimeError("No ZHVI rows parsed")
    dest = cstore.join(cache, "zillow_home_value_ts.csv")
    with cstore.open_write(dest, "w", newline="") as f:
        wr = csv.writer(f)
        wr.writerow(["GEOID", "NAME"] + [f"y{y}" for y in years])
        for fips in sorted(series):
            wr.writerow(
                [f"0500000US{fips}", ""]
                + [series[fips].get(y, "") for y in years]
            )
    logger.info("ZHVI: %d counties x %d years -> %s", len(series), len(years), dest)
    return _file_info(dest, ZHVI_URL, False)
